from dataclasses import *
from typing import List, Optional, Tuple
from openpyxl import load_workbook

question_time = 60
time_low_threshold = 5


@dataclass
class Game:
    categories: List['Category']
    length: int

    def get_category_by_name(self, name: str) -> Optional['Category']:
        """
        gets category by name if it exists
        :param name: cetegory or None
        :return:
        """
        for category in self.categories:
            if category.name == name:
                return category
        return None


@dataclass
class Category:
    name: str
    questions: List['Question']
    was_used: bool
    media_path: str

    def get_total_points(self) -> int:
        """
        returns total points for category
        :return: points
        """
        total = 0
        for question in self.questions:
            total += question.points

        return total

    def get_question_number(self) -> int:
        """
        returns total number of questions for category
        :return: number of questions
        """
        return len(self.questions)


@dataclass
class Question:
    category: str
    description: str
    points: int
    number: int
    answer: str
    show_answer: bool = False
    filepath: str = ""
    answer_filepath: str = ""


def create_game(filename: str) -> Tuple[Optional[Game], str]:
    """
    reads game questions from xlsx datafile
    :param filename:
    :return:
    """
    try:
        wb = load_workbook(filename=filename)
        sheet = wb.active
    except FileNotFoundError:
        return None, "file not found"
    # noinspection PyArgumentList
    game = Game(categories=list(), length=0)
    error = ""
    current_category = None
    for i in range(2, sheet.max_row + 1):
        try:
            category = sheet['A%i' % i].value.lower() if sheet['A%i' % i] and sheet['A%i' % i].value else ""
            number = int(sheet['B%i' % i].value) if sheet['B%i' % i] and sheet['B%i' % i].value else -1
            points = int(sheet['C%i' % i].value) if sheet['C%i' % i] and sheet['C%i' % i].value else -1
            text = sheet['D%i' % i].value if sheet['D%i' % i] and sheet['D%i' % i].value else ""
            filepath = sheet['E%i' % i].value if sheet['E%i' % i] and sheet['E%i' % i].value else ""
            answer = sheet['F%i' % i].value if sheet['F%i' % i] and sheet['F%i' % i].value else ""
            show_answer = True if sheet['G%i' % i] and sheet['G%i' % i].value and \
                                  sheet['G%i' % i].value.strip().lower() == 'да' else False
            answer_filepath = sheet['H%i' % i].value if sheet['H%i' % i] and sheet['H%i' % i].value else ""
            if category:
                existing_category = game.get_category_by_name(category)
                if not existing_category:
                    # noinspection PyArgumentList
                    existing_category = Category(name=category.lower(), questions=list(), was_used=False, media_path="")
                    game.categories.append(existing_category)
                    game.length += 1
                current_category = existing_category
            if (text or filepath) and points != -1:
                if current_category:
                    if number == -1:
                        number = len(current_category.questions) + 1
                    # noinspection PyArgumentList
                    question = Question(category=category.lower(), description=text, points=points, number=number,
                                        answer=answer, filepath=filepath, show_answer=show_answer,
                                        answer_filepath=answer_filepath)
                    current_category.questions.append(question)
                else:
                    error += "%i строчка без категории\n" % i
            elif category and points == -1 and answer == "" and number == -1:
                pass
            elif not category and not text and not filepath and not answer and points == number == -1:
                pass
            else:
                error += "%i строчка заполнена некорректно\n" % i
        except (AttributeError, TypeError, ValueError) as e:
            error += "Ошибка в  %i строчке\n" % i
    return game, error
